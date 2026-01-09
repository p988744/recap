"""Excel 匯出服務 - 產生績效考核表 Excel 檔案"""

import os
import logging
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from .db_service import add_title

logger = logging.getLogger(__name__)

# 預設模板路徑（套件內）
DEFAULT_TEMPLATE_PATH = Path(__file__).parent.parent / "data" / "template.xlsx"

# 常數定義
WORK_RESULTS_START_ROW = 3
SKILL_DEVELOPMENT_START_ROW = 10
SKILL_DEVELOPMENT_MAX_ITEMS = 4
ETHICS_ROW_MAP = {"責任感": 19, "團隊精神": 20, "主動積極": 21}
MAX_WORK_ITEMS = 10
DEFAULT_WEIGHT = 0.1


def export_to_excel(
    emp_name: str,
    emp_dept: str,
    emp_title: str,
    emp_start_date: str,
    emp_manager: str,
    emp_period: str,
    emp_sick_leave: float,
    emp_personal_leave: float,
    emp_absent: float,
    work_draft: str,
    skill_draft: str,
    ethics_draft: str,
    analysis_data: dict,
    template_path: str | Path | None = None
) -> str | None:
    """
    依照 template.xlsx 格式匯出績效考核表

    Returns:
        str: 產生的 Excel 檔案路徑，失敗時回傳 None
    """
    if not analysis_data:
        logger.warning("No analysis data available for export")
        return None

    # 儲存新的職稱到後端
    if emp_title:
        add_title(emp_title)

    # 使用 UI 輸入的員工姓名，若無則從 Tempo 資料取得
    user_name = emp_name or analysis_data.get('user_name', 'unknown')

    # 使用預設模板路徑
    if template_path is None:
        template_path = DEFAULT_TEMPLATE_PATH

    # 複製模板
    temp_file = tempfile.NamedTemporaryFile(
        delete=False, suffix='.xlsx', prefix=f'PE_{user_name}_'
    )
    shutil.copy(str(template_path), temp_file.name)
    logger.info(f"Created temp file: {temp_file.name}")

    try:
        # 載入複製的檔案
        wb = load_workbook(temp_file.name)

        # ===== 更新「績效檢討」工作表 =====
        ws_review = wb['績效檢討']
        ws_review['B1'] = user_name
        ws_review['E1'] = datetime.now().strftime('%Y-%m-%d')
        ws_review['B2'] = emp_dept
        ws_review['E2'] = emp_title
        ws_review['B3'] = emp_start_date
        ws_review['E3'] = emp_manager
        ws_review['B4'] = emp_period
        ws_review['A26'] = emp_sick_leave
        ws_review['B26'] = emp_personal_leave
        ws_review['C26'] = emp_absent

        # ===== 更新「工作成果」工作表 =====
        ws_work = wb['工作成果']

        work_items = work_draft.split("\n---\n")
        for idx, item_text in enumerate(work_items[:MAX_WORK_ITEMS], start=1):
            lines = item_text.strip().split("\n")
            if len(lines) < 3:
                continue

            title_line = lines[0]
            name = title_line.split("】")[1].strip() if "】" in title_line else f"工作項目 {idx}"

            weight_line = lines[1] if len(lines) > 1 else ""
            weight = DEFAULT_WEIGHT
            if "權重：" in weight_line:
                try:
                    weight_str = weight_line.split("權重：")[1].split("%")[0]
                    weight = float(weight_str) / 100
                except (IndexError, ValueError) as e:
                    logger.warning(f"Failed to parse weight from '{weight_line}': {e}")

            # 解析期間（如果有的話）
            item_period_start = ""
            item_period_end = ""
            details_start_line = 2  # 預設從第 3 行開始是成果說明

            for i, line in enumerate(lines[1:], start=1):
                if line.startswith("期間："):
                    period_str = line.replace("期間：", "").strip()
                    if "~" in period_str:
                        parts = period_str.split("~")
                        item_period_start = parts[0].strip() if len(parts) > 0 else ""
                        item_period_end = parts[1].strip() if len(parts) > 1 else ""
                    details_start_line = i + 1
                    break

            # 如果沒有解析到期間，使用整體考核期間
            if not item_period_start or not item_period_end:
                period_parts = emp_period.split('~') if '~' in emp_period else ['2025/01/01', '2025/12/31']
                item_period_start = period_parts[0].strip() if len(period_parts) > 0 else '2025/01/01'
                item_period_end = period_parts[1].strip() if len(period_parts) > 1 else '2025/12/31'

            # 過濾空行後取得成果說明
            details_lines = [l for l in lines[details_start_line:] if l.strip()]
            details = "\n".join(details_lines)

            row = idx + WORK_RESULTS_START_ROW - 1
            ws_work.cell(row=row, column=1, value=idx)
            ws_work.cell(row=row, column=2, value=name)
            ws_work.cell(row=row, column=3, value=f"{item_period_start}\n{item_period_end}")
            ws_work.cell(row=row, column=4, value=details)
            ws_work.cell(row=row, column=5, value=round(weight, 2))
            ws_work.cell(row=row, column=6, value=0)
            ws_work.cell(row=row, column=8, value=0)

        # ===== 更新技能發展 =====
        skill_items = []
        current_skill = None
        for line in skill_draft.split("\n"):
            if line.startswith("【") and "】" in line:
                if current_skill:
                    skill_items.append(current_skill)
                skill_name = line.split("】")[0].replace("【", "").replace("技能項目", "").strip()
                if not skill_name:
                    skill_name = line.split("】")[1].strip() if len(line.split("】")) > 1 else ""
                current_skill = {"name": skill_name, "desc": ""}
            elif current_skill and line.strip():
                if line.startswith("具體說明："):
                    current_skill["desc"] = line.replace("具體說明：", "").strip()
                else:
                    current_skill["desc"] += " " + line.strip()
        if current_skill:
            skill_items.append(current_skill)

        for idx, skill in enumerate(skill_items[:SKILL_DEVELOPMENT_MAX_ITEMS], start=1):
            row = SKILL_DEVELOPMENT_START_ROW - 1 + idx
            ws_review.cell(row=row, column=2, value=skill.get("name", ""))
            ws_review.cell(row=row, column=6, value=skill.get("desc", ""))

        # ===== 更新職場專業素養 =====
        current_ethics = None
        ethics_items = {}

        for line in ethics_draft.split("\n"):
            for key in ETHICS_ROW_MAP:
                if key in line and "【" in line:
                    current_ethics = key
                    break
            if current_ethics and line.strip() and "【" not in line:
                if line.startswith("具體說明"):
                    ethics_items[current_ethics] = line.replace("具體說明：", "").replace("具體說明", "").strip()
                elif current_ethics in ethics_items:
                    ethics_items[current_ethics] += " " + line.strip()
                else:
                    ethics_items[current_ethics] = line.strip()

        for key, row in ETHICS_ROW_MAP.items():
            if key in ethics_items:
                ws_review.cell(row=row, column=6, value=ethics_items[key])

        wb.save(temp_file.name)
        wb.close()

        return temp_file.name

    except Exception as e:
        logger.exception(f"Error exporting to Excel: {e}")
        return None

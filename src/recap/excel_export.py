"""
Excel 報表匯出模組

將團隊工時報表匯出成 Excel 格式
"""

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .team import TeamReportData


class ExcelReportExporter:
    """Excel 報表匯出器"""

    def __init__(self, report: "TeamReportData"):
        """
        初始化匯出器

        Args:
            report: 團隊報表資料
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "Excel 匯出需要 openpyxl 套件。\n"
                "請執行: pip install recap[excel]"
            )

        self.report = report
        self.wb = Workbook()
        self._Font = Font
        self._PatternFill = PatternFill
        self._Alignment = Alignment
        self._Border = Border
        self._Side = Side
        self._get_column_letter = get_column_letter

        # 樣式定義
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        self.total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    def export(self, filepath: str):
        """
        匯出報表到 Excel 檔案

        Args:
            filepath: 輸出檔案路徑
        """
        # 移除預設的空白工作表
        self.wb.remove(self.wb.active)

        # 建立各個工作表
        self._create_summary_sheet()
        self._create_by_type_sheet()
        self._create_by_date_sheet()
        self._create_raw_data_sheet()

        # 儲存檔案
        self.wb.save(filepath)

    def _create_summary_sheet(self):
        """建立總覽工作表"""
        ws = self.wb.create_sheet("Summary")

        # 報表標題區
        ws["A1"] = "Team Worklog Report"
        ws["A1"].font = self._Font(bold=True, size=16)
        ws.merge_cells("A1:D1")

        ws["A3"] = "Team:"
        ws["B3"] = self.report.team_name
        ws["A4"] = "Jira Group:"
        ws["B4"] = self.report.jira_group
        ws["A5"] = "Period:"
        ws["B5"] = f"{self.report.start_date} ~ {self.report.end_date}"
        ws["A6"] = "Generated:"
        ws["B6"] = self.report.generated_at[:19].replace("T", " ")
        ws["A7"] = "Total Hours:"
        ws["B7"] = f"{self.report.total_hours:.1f}"
        ws["B7"].font = self._Font(bold=True)

        # 成員工時表
        issue_types = self.report.get_all_issue_types()
        start_row = 9

        # 表頭
        headers = ["Member", "Total Hours"] + issue_types
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 資料列
        for row_idx, member in enumerate(self.report.members, start_row + 1):
            # 成員名稱
            cell = ws.cell(row=row_idx, column=1, value=member.member.display_name)
            cell.alignment = self.left_align
            cell.border = self.thin_border

            # 總工時
            cell = ws.cell(row=row_idx, column=2, value=round(member.total_hours, 1))
            cell.alignment = self.center_align
            cell.border = self.thin_border

            # 各類型工時
            for col, issue_type in enumerate(issue_types, 3):
                hours = member.by_issue_type.get(issue_type, 0) / 3600
                cell = ws.cell(row=row_idx, column=col, value=round(hours, 1) if hours > 0 else "")
                cell.alignment = self.center_align
                cell.border = self.thin_border

        # 合計列
        total_row = start_row + len(self.report.members) + 1
        cell = ws.cell(row=total_row, column=1, value="Total")
        cell.font = self._Font(bold=True)
        cell.fill = self.total_fill
        cell.border = self.thin_border

        cell = ws.cell(row=total_row, column=2, value=round(self.report.total_hours, 1))
        cell.font = self._Font(bold=True)
        cell.fill = self.total_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border

        by_type = self.report.by_issue_type_total
        for col, issue_type in enumerate(issue_types, 3):
            cell = ws.cell(row=total_row, column=col, value=round(by_type.get(issue_type, 0), 1))
            cell.font = self._Font(bold=True)
            cell.fill = self.total_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 調整欄寬
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(issue_types) + 3):
            ws.column_dimensions[self._get_column_letter(col)].width = 12

    def _create_by_type_sheet(self):
        """建立依 Issue 類型分析工作表"""
        ws = self.wb.create_sheet("By Issue Type")

        issue_types = self.report.get_all_issue_types()

        # 表頭
        headers = ["Issue Type", "Total Hours"] + [m.member.display_name for m in self.report.members]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 資料列
        by_type = self.report.by_issue_type_total
        for row, issue_type in enumerate(issue_types, 2):
            # Issue 類型
            cell = ws.cell(row=row, column=1, value=issue_type)
            cell.alignment = self.left_align
            cell.border = self.thin_border

            # 總工時
            cell = ws.cell(row=row, column=2, value=round(by_type.get(issue_type, 0), 1))
            cell.alignment = self.center_align
            cell.border = self.thin_border

            # 各成員工時
            for col, member in enumerate(self.report.members, 3):
                hours = member.by_issue_type.get(issue_type, 0) / 3600
                cell = ws.cell(row=row, column=col, value=round(hours, 1) if hours > 0 else "")
                cell.alignment = self.center_align
                cell.border = self.thin_border

        # 調整欄寬
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(self.report.members) + 3):
            ws.column_dimensions[self._get_column_letter(col)].width = 15

    def _create_by_date_sheet(self):
        """建立依日期分析工作表"""
        ws = self.wb.create_sheet("By Date")

        dates = self.report.get_all_dates()

        # 表頭
        headers = ["Date", "Total Hours"] + [m.member.display_name for m in self.report.members]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 資料列
        by_date = self.report.by_date_total
        for row, date in enumerate(dates, 2):
            # 日期
            cell = ws.cell(row=row, column=1, value=date)
            cell.alignment = self.left_align
            cell.border = self.thin_border

            # 總工時
            cell = ws.cell(row=row, column=2, value=round(by_date.get(date, 0), 1))
            cell.alignment = self.center_align
            cell.border = self.thin_border

            # 各成員工時
            for col, member in enumerate(self.report.members, 3):
                hours = member.by_date.get(date, 0) / 3600
                cell = ws.cell(row=row, column=col, value=round(hours, 1) if hours > 0 else "")
                cell.alignment = self.center_align
                cell.border = self.thin_border

        # 合計列
        if dates:
            total_row = len(dates) + 2
            cell = ws.cell(row=total_row, column=1, value="Total")
            cell.font = self._Font(bold=True)
            cell.fill = self.total_fill
            cell.border = self.thin_border

            cell = ws.cell(row=total_row, column=2, value=round(self.report.total_hours, 1))
            cell.font = self._Font(bold=True)
            cell.fill = self.total_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

            for col, member in enumerate(self.report.members, 3):
                cell = ws.cell(row=total_row, column=col, value=round(member.total_hours, 1))
                cell.font = self._Font(bold=True)
                cell.fill = self.total_fill
                cell.alignment = self.center_align
                cell.border = self.thin_border

        # 調整欄寬
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        for col in range(3, len(self.report.members) + 3):
            ws.column_dimensions[self._get_column_letter(col)].width = 15

    def _create_raw_data_sheet(self):
        """建立原始資料工作表（供樞紐分析）"""
        ws = self.wb.create_sheet("Raw Data")

        # 表頭
        headers = ["Member", "Date", "Issue", "Issue Type", "Hours", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 資料列
        row = 2
        for member in self.report.members:
            for entry in member.entries:
                ws.cell(row=row, column=1, value=member.member.display_name)
                ws.cell(row=row, column=2, value=entry.date)
                ws.cell(row=row, column=3, value=entry.issue_key)
                ws.cell(row=row, column=4, value=entry.issue_type)
                ws.cell(row=row, column=5, value=round(entry.time_spent_seconds / 3600, 2))
                ws.cell(row=row, column=6, value=entry.description[:200] if entry.description else "")

                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.thin_border

                row += 1

        # 調整欄寬
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 50
